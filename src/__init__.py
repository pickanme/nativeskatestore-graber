# import pyodbc
from io import StringIO
from urllib.request import Request, urlopen
from lxml import etree
from openpyxl import Workbook

# Specifying the ODBC driver, server name, database, etc. directly
# cnxn = pyodbc.connect('DRIVER={/usr/local/lib/libsqlite3odbc.so};'
#                       'Database=../resources/nativeskatestore.sqlite;'
#                       'LongNames=0;Timeout=1000;NoTXN=0;'
#                       'SyncPragma=NORMAL;StepAPI=0;')
# # Python 3.x
# cnxn.setdecoding(pyodbc.SQL_WCHAR, encoding='utf-8')
# cnxn.setencoding(encoding='utf-8')
# # Create a cursor from the connection
# cursor = cnxn.cursor()

# qs_x = "//div[contains(@class, 'qa-q-list')]/div[contains(@class, 'qa-q-list-item')]"
next_page = 'https://i-otvet.ru/questions'
cat_x = "//span[contains(@class, 'qa-q-item-where-data')]/a[contains(@class, 'qa-category-link')]"
categories = set()

def main():
    print_products(next_page)
    for i in range(20,10952380,20):
        print_products(next_page + "?start=" + str(i))

    i = 1
    wb = Workbook()
    ws = wb.active

    for cat in categories:
        print("Add ", i, " category")
        ws.cell(row=i, column=1, value=i)
        ws.cell(row=i, column=2, value=cat)
        i = i + 1

    wb.save(filename = "cats.xls")


def print_products(link):
    req = Request(link, headers={'User-Agent': 'Mozilla/5.0'})
    data = urlopen(req).read()  # bytes
    body = data.decode('utf-8')

    parser = etree.HTMLParser()
    tree = etree.parse(StringIO(body), parser)

    cat_names = tree.xpath(cat_x)

    for cat in cat_names:
        categories.add(cat.text)

    print(link, " is scanned.")

if __name__ == '__main__':
    main()
