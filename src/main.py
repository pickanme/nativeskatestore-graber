#!/usr/bin/python
# -*- coding: utf-8 -*-

import os
import logging
from io import StringIO
from urllib.request import Request, urlopen
from lxml import etree
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime
from classes.answer import Answer
from classes.question import Question

# create logger
cli_logger = logging.getLogger('fonbet-parser')
# set logging level
cli_logger.setLevel(logging.DEBUG)
# create console and file handler
ch = logging.StreamHandler()
# create formatter
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
# add formatter to handlers
ch.setFormatter(formatter)
# add handlers to logger
cli_logger.addHandler(ch)

root = "c:/Users/Dinir/Documents/git/nativeskatestore-graber/"
q_file = "resources/sample-file.xlsx"
cat_file = "resources/categories.xlsx"

if os.path.isfile(q_file):
    wb_q = load_workbook(filename=os.path.join(root, q_file))
    ID_COL = 1
    TYPE_COL = 2
    PARENT_ID_COL = 3
    TITLE_COL = 5
    CONTENT_COL = 6
    FORMAT_COL = 7
    CATEGORY_COL = 8
    TAGS_COL = 10
    USERNAME_COL = 12
    DATETIME_COL = 15
else:
    wb_q = Workbook()
if os.path.isfile(cat_file):
    wb_cat = load_workbook(filename=os.path.join(root, cat_file))
else:
    wb_cat = Workbook()

q_id = 1
c_id = 1
last_free_row = 2

qs_x = "//div[contains(@class, 'qa-q-list')]/div[contains(@class, 'qa-q-list-item')]/div[contains(@class, 'qa-q-item-main')]/div[contains(@class, 'qa-q-item-title')]/a"
q_block_x = "//div[contains(@class, 'qa-part-q-view')]//div[contains(@class, 'qa-q-view-main')]"
q_title_x = "//div[contains(@class, 'qa-main-heading')]/h1/a/span"
q_content_x = "./form/div[contains(@class, 'qa-q-view-content')]/div[contains(@itemprop, 'text')]"
q_tags_x = "./form/div[contains(@class, 'qa-q-view-tags')]" \
            "/ul[contains(@class, 'qa-q-view-tag-list')]" \
            "/li[contains(@class, 'qa-q-view-tag-item')]/a"
q_date_x = "./form/span[contains(@class, 'qa-q-view-avatar-meta')]/span" \
            "/span[contains(@class, 'qa-q-view-when')]" \
            "/span[contains(@class, 'qa-q-view-when-data')]/time"
q_username_x = "./form/span[contains(@class, 'qa-q-view-avatar-meta')]/span" \
            "/span[contains(@class, 'qa-q-view-who')]" \
            "/span[contains(@class, 'qa-q-view-who-data')]/span/a/span[contains(@itemprop, 'name')]"
q_category_x = "./form/span[contains(@class, 'qa-q-view-avatar-meta')]/span" \
            "/span[contains(@class, 'qa-q-view-where')]" \
            "/span[contains(@class, 'qa-q-view-where-data')]/a"

answers_x = "//div[contains(@class, 'qa-part-a-list')]/div[contains(@class, 'qa-a-list')]/div[contains(@class, 'qa-a-list-item')]/div[contains(@class, 'qa-a-item-main')]"
ans_next_page_x = "//li[contains(@class, 'qa-page-links-item')]/a[contains(@class, 'qa-page-next')]"

ans_content_x = "./form/div[contains(@class, 'qa-a-item-content')]/div[contains(@itemprop, 'text')]"
ans_date_x = "./form//span[contains(@class, 'qa-a-item-avatar-meta')]/span" \
            "/span[contains(@class, 'qa-a-item-when')]" \
            "/span[contains(@class, 'qa-a-item-when-data')]/time"
ans_username_x = "./form//span[contains(@class, 'qa-a-item-avatar-meta')]/span" \
            "/span[contains(@class, 'qa-a-item-who')]" \
            "/span[contains(@class, 'qa-a-item-who-data')]/span/a/span[contains(@itemprop, 'name')]"

date_p = "%Y-%m-%dT%H:%M:%S%z" 
date_f = "%d.%m.%Y %H:%M:%S"

host = "https://i-otvet.ru"
next_page = host+"/questions"

q_list = []
a_list = []
c_dict = dict()

def main():
    process_list_page(next_page)
    for i in range(20,10952381,20):
        process_list_page(next_page + "?start=" + str(i))
        if (i % 100 == 0):
            cli_logger.warn("Writing in " + cat_file)
            add_cats()
            cli_logger.warn("Writing in " + q_file)
            add_ques()
            add_ans()
            q_list = []
            a_list = []
            if (os.stat(root+q_file).st_size < 1048576):
                pass


def process_list_page(l_link):
    tree = get_tree(l_link)
    # get question elements
    questions = tree.xpath(qs_x)
    # get links from elements
    for q in questions:
        que_link = host+q.attrib.get('href')[1:]
        process_q_page(que_link)
    
    cli_logger.info(l_link)

def process_q_page(q_link):
    global q_id
    # print(q_link)
    tree = get_tree(q_link)

    # grab answer title
    title = tree.xpath(q_title_x)[0].text

    # grab question block
    q_block = tree.xpath(q_block_x)[0]
    # grab q_block content
    content_el = q_block.xpath(q_content_x)
    content = '' 
    if(len(content_el) != 0):
        content = stringify_children(content_el[0])
    # grab q_block datetime 
    datetime_from = get_date(q_block.xpath(q_date_x)[0].attrib.get('datetime'))
    # grab q_block tags
    tags = get_tags(q_block.xpath(q_tags_x))
    # grab q_block username
    user_el = q_block.xpath(q_username_x)
    username = 'аноним'
    if(len(user_el) != 0):
        username = user_el[0].text
    # grab q_block category
    cat_el = q_block.xpath(q_category_x)
    category_id = ''
    if (len(cat_el) != 0):
        category_id = get_id(cat_el[0].text)

    q_list.append(Question(q_id, tags, title, username, datetime_from, content, category_id))
    process_answers(q_id, tree)
    
    q_id = q_id + 1


def add_ques():
    global last_free_row

    wb_q_s = wb_q.active
    
    for q in q_list:
        wb_q_s.cell(column=ID_COL, row=last_free_row).value = q.q_id
        wb_q_s.cell(column=TYPE_COL, row=last_free_row).value = "Q"
        wb_q_s.cell(column=TITLE_COL, row=last_free_row).value = q.title
        wb_q_s.cell(column=CONTENT_COL, row=last_free_row).value = q.content
        wb_q_s.cell(column=FORMAT_COL, row=last_free_row).value = "html"
        wb_q_s.cell(column=CATEGORY_COL, row=last_free_row).value = q.cat_id
        wb_q_s.cell(column=TAGS_COL, row=last_free_row).value = q.tags
        wb_q_s.cell(column=USERNAME_COL, row=last_free_row).value = q.username
        wb_q_s.cell(column=DATETIME_COL, row=last_free_row).value = q.datetime_from
        last_free_row = last_free_row + 1
    
    wb_q.save(filename=q_file)

    
def process_answers(p_id, tree):
    answers = tree.xpath(answers_x)
    # print(len(answers))
    # print("Processing answers")
    for ans in answers:
        # grab ans content
        content_el = ans.xpath(ans_content_x)
        content = '' 
        if(len(content_el) != 0):
            content = stringify_children(content_el[0])
        # grab ans datetime 
        datetime_from = get_date(ans.xpath(ans_date_x)[0].attrib.get('datetime'))
        # grab ans username
        user_el = ans.xpath(ans_username_x)
        username = 'аноним'
        if(len(user_el) != 0):
            username = user_el[0].text

        a_list.append(Answer(p_id, username, datetime_from, content))
        # add to excel doc
        # add_ans(p_id, username, datetime, content)
    
    np_link = tree.xpath(ans_next_page_x)
    if(len(np_link) != 0):
        an_page = host+np_link.attrib.get('href')[1:]
        process_answers(p_id, an_page)

def add_ans():
    global last_free_row
    
    wb_q_s = wb_q.active

    for a in a_list:
        wb_q_s.cell(column=PARENT_ID_COL, row=last_free_row).value = a.p_id
        wb_q_s.cell(column=TYPE_COL, row=last_free_row).value = "A"
        wb_q_s.cell(column=CONTENT_COL, row=last_free_row).value = a.content
        wb_q_s.cell(column=FORMAT_COL, row=last_free_row).value = "html"
        wb_q_s.cell(column=USERNAME_COL, row=last_free_row).value = a.username
        wb_q_s.cell(column=DATETIME_COL, row=last_free_row).value = a.datetime_from
        last_free_row = last_free_row + 1
    
    wb_q.save(filename=q_file)

def add_cats():
    wb_cat_s = wb_cat.active
    
    for c_name, c_id in c_dict.items():
        wb_cat_s.cell(column=1, row=c_id+1).value = c_id
        wb_cat_s.cell(column=2, row=c_id+1).value = c_name

    wb_cat.save(filename=cat_file)


def get_tree(link):
    req = Request(link, headers={'User-Agent': 'Mozilla/5.0'})
    data = urlopen(req).read()  # bytes
    body = data.decode('utf-8')

    parser = etree.HTMLParser()
    tree = etree.parse(StringIO(body), parser)

    return tree

def get_date(date_str):
    # parse string
    date_obj = datetime.strptime(date_str, date_p)
    # set format for a date object
    return datetime.strftime(date_obj, date_f)

def get_tags(tags):
    t_str = ""
    for tag in tags:
        t_str += tag.text+", "

    return t_str[0:-2]

def get_id(c_name):
    return c_dict[c_name] if c_name in c_dict else create_cat(c_name)

def create_cat(c_name):
    global c_id
    c_id = c_id + 1
    c_dict[c_name] = c_id

    return c_id

# https://stackoverflow.com/questions/4624062/get-all-text-inside-a-tag-in-lxml
def stringify_children(node):
    from lxml.etree import tostring
    from itertools import chain
    parts = ([node.text] +
            list(chain(*([c.text, tostring(c, encoding=str), c.tail] for c in node.getchildren()))) +
            [node.tail])
    # filter removes possible Nones in texts and tails
    return ''.join(filter(None, parts))

if __name__ == '__main__':
	main()